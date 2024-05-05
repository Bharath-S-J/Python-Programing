import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # Load the workbook and select the target sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    
    # Example of accessing a specific cell by coordinates or name
    cell_a1 = sheet['A1']
    cell_1_1 = sheet.cell(1, 1)

    # Print the value of the cell (e.g., A1)
    print(cell_a1.value)

    # Determine the maximum number of rows in the sheet
    max_rows = sheet.max_row
    print("Max Rows:", max_rows)
    print('---')

    # Iterate through each row starting from row 2 (skipping header)
    for row in range(2, max_rows + 1):
        # Access the cell in the third column (C) for each row
        cell = sheet.cell(row, 3)
        print("Original Price (Row {}): {}".format(row, cell.value))

        # Calculate the corrected price (e.g., 90% of original price)
        corrected_price = cell.value * 0.9

        # Update the value of the cell in the fourth column (D) with corrected price
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Define the data range (values) for creating the bar chart
    values = Reference(sheet, min_row=2, max_row=max_rows, min_col=4, max_col=4)

    # Create a BarChart object and add data values
    chart = BarChart()
    chart.add_data(values)

    # Add the chart to a specific location (e.g., cell 'E2') in the sheet
    sheet.add_chart(chart, 'E2')

    # Save the modified workbook to a new file (or overwrite the existing one)
    wb.save('transactions2.xlsx')

# Example usage: Process a specific workbook file ('transactions2.xlsx')
process_workbook('transactions.xlsx')
